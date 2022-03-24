from openpyxl import load_workbook


def insert_sticker(keyword, sticker_id=None, reply_text=None):
    row = stickers_page.max_row + 1
    stickers_page.cell(row=row, column=1).value = keyword
    stickers_page.cell(row=row, column=2).value = sticker_id
    stickers_page.cell(row=row, column=3).value = reply_text
    bd.save('database.xlsx')

    stickers[keyword] = sticker_id
    replies[keyword] = reply_text


def insert_user(user_id, name, sex, grade):
    '''
    вносит нового пользователя в базу данных
    '''
    user_page = bd['Users']
    row = user_page.max_row + 1
    user_page.cell(row=row, column=1).value = user_id
    user_page.cell(row=row, column=2).value = name
    user_page.cell(row=row, column=3).value = sex
    user_page.cell(row=row, column=3).value = grade
    bd.save('database.xlsx')


def in_database(user: int) -> bool:
    '''
    возвращает True, если id пользователь есть в database
    '''
    user_page = bd['Users']
    for row in range(1, user_page.max_row + 1):
        if user == user_page.cell(row=row, column=1).value:
            return True
        return False


bd = load_workbook('database.xlsx')
for sheet in bd:
    print(sheet.title)
stickers_page = bd['stickers']


stickers = {}
replies = {}

for row in range(1, stickers_page.max_row + 1):
    keyword = stickers_page.cell(row=row, column=1).value
    sticker_id = stickers_page.cell(row=row, column=2).value
    reply_text = stickers_page.cell(row=row, column=3).value
    stickers[keyword] = sticker_id
    replies[keyword] = reply_text


if __name__ == '__main__':
    insert_sticker('жиза', reply_text='жизнь - интересная штука')
    print(stickers)
