from openpyxl import load_workbook

bd = load_workbook('EXEL.xlsx')
for sheet in bd:
    print(sheet.title)
stickers_page = bd['stickers']


stickers = {}


for row in range(1, stickers_page.max_row + 1):
    keyword = stickers_page.cell(row=row, column=1).value
    sticker_id = stickers_page.cell(row=row, column=2).value
    print(keyword, sticker_id)
    stickers[keyword] = sticker_id

print(stickers)

'''

     if stickers_page.cell(row=row, column=column).value == 'пока':
            print(stickers_page.cell(row=row, column=column + 1).value)
        elif stickers_page.cell(row=row, column=column).value == 'как дела?':
            print(stickers_page.cell(row=row, column=column + 1).value)
        elif stickers_page.cell(row=row, column=column).value == 'удачи':
            print(stickers_page.cell(row=row, column=column + 1).value)
        elif stickers_page.cell(row=row, column=column).value == 'знаешь?':
            print(stickers_page.cell(row=row, column=column + 1).value)
            '''

